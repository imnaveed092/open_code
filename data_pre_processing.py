from openpyxl import load_workbook

def handle_spilled_data_and_delete_spilled_row(file_path, sheet_name):
    """
    Handles spilled data by concatenating spilled data into the original row and deleting the spilled row.
    
    :param file_path: Path to the Excel file.
    :param sheet_name: Name of the worksheet to process.
    """
    try:
        # Load the workbook and select the sheet
        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]

        max_row = sheet.max_row
        max_column = sheet.max_column  # Get the actual number of columns (which may be larger than V)

        row = 1
        while row <= max_row:
            # Check if the current row has spilled data (next row is empty, next next has data)
            next_row_empty = all(sheet.cell(row=row + 1, column=col).value is None for col in range(1, max_column + 1))
            next_next_row_data = any(sheet.cell(row=row + 2, column=col).value is not None for col in range(1, max_column + 1))

            # If there is spilled data (next row is empty, and the row after that has data)
            if next_row_empty and next_next_row_data:
                # Collect spilled data from the next row (e.g., row 199)
                spilled_data = []
                # We iterate over all columns up to max_column to account for spills beyond column V
                for col in range(1, max_column + 1):
                    cell_value = sheet.cell(row=row + 2, column=col).value
                    if cell_value:  # Only add non-empty cells
                        spilled_data.append(str(cell_value))

                # Concatenate the spilled data into a single string
                concatenated_data = " ".join(spilled_data)

                # Find the next available column in the original row (after column V)
                target_col = 22  # Column V is the 22nd column (A=1, B=2, ..., V=22)
                while sheet.cell(row=row, column=target_col).value is not None:
                    target_col += 1

                # Place the concatenated spilled data in the original row (row 197)
                sheet.cell(row=row, column=target_col).value = concatenated_data

                # Delete the row where the spilled data was (row 199)
                sheet.delete_rows(row + 2)  # Row + 2 because spilled data is in row + 2

                # Recalculate max_row since we've deleted a row
                max_row -= 1

            # Move to the next row (after processing any possible spill)
            row += 1

        # Save the updated file
        workbook.save(file_path)
        print(f"Spilled data has been merged into the original row, and the spilled row has been deleted.")

    except Exception as e:
        print(f"An error occurred: {e}")


# Example usage
handle_spilled_data_and_delete_spilled_row("./New Microsoft Excel Worksheet.xlsx", "Sheet1")
